import json
import os
from typing import List, Optional

from openpyxl import Workbook, load_workbook


def manipulate_xlsx(
    operation: str,
    filename: Optional[str] = None,
    sheet_name: str = "Sheet1",
    data: Optional[List[List[str]]] = None,
    header: bool = False,
) -> str:
    """
    Manipula archivos Excel (.xlsx) en el directorio raíz.

    Args:
        operation: 'read', 'write', 'append' o 'list_files'
        filename: Nombre del archivo (no requerido para list_files)
        sheet_name: Nombre de la hoja (default: 'Sheet1')
        data: Datos a escribir (para write/append)
        header: Indica si los datos incluyen encabezados

    Returns:
        str: JSON con los resultados o mensaje de error
    """
    try:
        if operation == "list_files":
            xlsx_files = [f for f in os.listdir() if f.endswith(".xlsx")]
            return json.dumps({"status": "success", "files": xlsx_files})

        if not filename:
            return json.dumps(
                {
                    "status": "error",
                    "message": "Se requiere un nombre de archivo para esta operación",
                }
            )

        filepath = os.path.join(os.getcwd(), filename)

        if operation == "read":
            if not os.path.exists(filepath):
                return json.dumps(
                    {"status": "error", "message": f"Archivo no encontrado: {filename}"}
                )

            wb = load_workbook(filepath)
            sheet = wb[sheet_name]

            result = []
            for row in sheet.iter_rows(values_only=True):
                result.append(list(row))

            return json.dumps({"status": "success", "data": result})

        elif operation in ["write", "append"]:
            if not data:
                return json.dumps(
                    {
                        "status": "error",
                        "message": "Se requieren datos para esta operación",
                    }
                )

            if operation == "write":
                wb = Workbook()
                sheet = wb.active
                sheet.title = sheet_name
            else:  # append
                if not os.path.exists(filepath):
                    return json.dumps(
                        {
                            "status": "error",
                            "message": f"Archivo no encontrado para añadir datos: {filename}",
                        }
                    )
                wb = load_workbook(filepath)
                if sheet_name not in wb.sheetnames:
                    sheet = wb.create_sheet(sheet_name)
                else:
                    sheet = wb[sheet_name]

            # Escribir datos
            for row in data:
                sheet.append(row)

            wb.save(filepath)
            return json.dumps(
                {
                    "status": "success",
                    "message": f"Archivo {filename} {'creado' if operation == 'write' else 'actualizado'} exitosamente",
                }
            )

        else:
            return json.dumps(
                {"status": "error", "message": f"Operación no válida: {operation}"}
            )

    except Exception as e:
        return json.dumps(
            {
                "status": "error",
                "message": f"Error al manipular archivo Excel: {str(e)}",
            }
        )


async def async_manipulate_xlsx(*args, **kwargs) -> str:
    return manipulate_xlsx(*args, **kwargs)
